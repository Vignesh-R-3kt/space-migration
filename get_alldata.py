import datetime
import threading
import requests
import openpyxl

import threading
import queue

# Create a queue to store the XLSX file names
xlsx_queue = queue.Queue()

# Set your Confluence server URL and API token
source_base_url = "https://inpixon.atlassian.net/wiki"
target_base_url = "https://cxapplucy.atlassian.net/wiki"
username = "mail6gowthamdwd@gmail.com"
password = "ATATT3xFfGF0as-kz6g-kLbLRj2NtiSYv6Mt4d_JNJicl92zPoY0cGaTIu2t5qcPjMgPCvADotB7pz7Yds5kYFj5bXxKuGpnofZqjQAPxZG1k2dtSVQnZKnZKOgfM5NpRYPuLiq5cDAXeIJxuYVJ8KTIGdtQJI2KY-4LfSLzGUvEnUfxq5TaOzc=F9B0A2B9"

# Set the space key for the space you want to retrieve pages from
space_key = "DATA"

# Initialize variables

start_val = 0
page_size = 10000  # Adjust this based on your needs
headers = {
    "Content-Type": "application/json",
}


# wb_matched = openpyxl.Workbook()
# ws_matched = wb_matched.active
# ws_matched.append(["timestamp", "Meta_data", "SourceID", "TargetID", 
#                    "Title", "size", "comment", 
#                    'sourceDisplayName', 'targetDisplayName',
#                    'sourcePublicName', 'targetPublicName', 
#                    'sourceDate', 'targetDate'])

# wb_missing = openpyxl.Workbook()
# ws_missing = wb_missing.active
# ws_missing.append(["timestamp", "Meta_data", "SourceID", "TargetID", 
#                    "Title", "size", "comment", 
#                    'sourceDisplayName', 'targetDisplayName',
#                    'sourcePublicName', 'targetPublicName', 
#                    'sourceDate', 'targetDate'])


def get_page_history(url, page_id):
    history_response = requests.get(f"{url}/rest/api/content/{page_id}/history",
                             headers=headers, auth=(username, password))
    creation_details = []
    if history_response.status_code == 200:
        date = history_response.json()['createdDate']
        display_name = history_response.json()['createdBy']['displayName']
        public_name = history_response.json()['createdBy']['publicName']
        return [display_name, public_name, date]


def get_page_comments(url, page_id):
    response = requests.get(f"{url}/rest/api/content/{page_id}/child/comment",
                             headers=headers, auth=(username, password))
    if response.status_code == 200:
        list_comments = response.json()['results']
        for comment in list_comments:
            return comment['title']

def validate_attachements(url, page_id):
    response = requests.get(f"{url}/rest/api/content/{page_id}/child/attachment",
                             headers=headers, auth=(username, password))
    
    if response.status_code == 200:
        data = response.json()
        list_attachments = data.get('results')
        page_attachments = {}
        for attachment in list_attachments:
            page_attachments.update({attachment['title']: 
                                     [attachment['id'], 
                                      attachment['extensions']['fileSize']]
                                    }
                            )

        return page_attachments

def get_target_pages(url, start_val, space_key):
    all_pages = []
    while True:
        # Define the API endpoint to retrieve pages in a space
        api_endpoint = f"{url}/rest/api/content?spaceKey={space_key}&start={start_val}&limit={page_size}"

        # Set up headers for authentication

        # Make the API request
        response = requests.get(api_endpoint, headers=headers, auth=(username, password))

        if response.status_code == 200:
            # Parse the JSON response and append pages to the result list
            page_data = response.json()
            all_pages.extend(page_data.get("results", []))

            # Check if there are more pages to fetch
            if not page_data.get("isLastPage", True):
                start_val += page_size
            else:
                break
        else:
            print(f"Failed to retrieve pages. Status code: {response.status_code}")
            break
    return all_pages


def initiate_space_migration(space_key, ws_matched, ws_missing):
    mapped_pageid = {}
    page_target_attachments = {}
    target_page_history = {}
    target_page_history_att = {}
    target_pages = get_target_pages(target_base_url, start_val, space_key)

    for page in target_pages:
        mapped_pageid.update({page['title']: page['id']})
        list_target_page_history = get_page_history(target_base_url,  page['id'])
        target_page_history.update({page['title']: list_target_page_history})
        target_attachments = validate_attachements(target_base_url, page['id'])
        for title, id in target_attachments.items():
                lis_tar_att = get_page_history(target_base_url,  id[0])
                target_page_history_att.update({page['title']: lis_tar_att})
        if target_attachments:
            page_target_attachments.update(target_attachments)


    page_source_attachments = {}
    source_pages = get_target_pages(source_base_url, start_val, space_key)

    for page in source_pages:
        page_title = page['title']
        lis_sorc = get_page_history(source_base_url,  page['id'])
        if not lis_sorc:
            lis_sorc = ["", "", ""]

        target_page_id = mapped_pageid.get(page_title, "")
        comment = get_page_comments(source_base_url, page['id'])
        lis_tar = target_page_history.get(page['title'])
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if not lis_tar:
            lis_tar = ["", "", ""]
        if target_page_id:
            ws_matched.append([timestamp, "page", page['id'], target_page_id, 
                            page_title, "", comment, lis_sorc[0],
                            lis_tar[0], lis_sorc[1], lis_tar[1], 
                            lis_sorc[2], lis_tar[1]])
        else:
            ws_missing.append([timestamp, "page", page['id'], target_page_id, 
                            page_title, "", comment, lis_sorc[0],
                            lis_tar[0], lis_sorc[1], lis_tar[1], 
                            lis_sorc[2], lis_tar[1]])
        source_attachments = validate_attachements(source_base_url, page['id'])
        if source_attachments:
            page_source_attachments.update(source_attachments)
            lis_tar_att = target_page_history_att.get(page['title'])
            for title, id in source_attachments.items():
                lis_sorc_att = get_page_history(source_base_url,  id[0])
                if not lis_sorc_att:
                    lis_sorc_att = ["", "", ""]
                if title in page_target_attachments:
                    target_id = page_target_attachments.get(title)[0]
                    img_size = page_target_attachments.get(title)[1]
                    
                    if not lis_tar_att:
                        lis_tar_att = ["", "", ""]
                    ws_matched.append([timestamp, "attachment", id[0], target_id, 
                                    title, id[1],"",
                                    lis_sorc_att[0], lis_tar_att[0],
                                    lis_sorc_att[1], lis_tar_att[1],
                                    lis_sorc_att[2], lis_tar_att[2]])
                else:
                    ws_missing.append([timestamp, "attachment", id[0], target_id, 
                                    title, id[1],"",
                                    lis_sorc_att[0], lis_tar_att[0],
                                    lis_sorc_att[1], lis_tar_att[1],
                                    lis_sorc_att[2], lis_tar_att[2]])



wb_matched_lock = threading.Lock()
wb_missing_lock = threading.Lock()

def process_space(space_key):
    # Initialize workbooks and worksheets for each space
    list_matchedData = []
    list_missingData = []
    wb_matched = openpyxl.Workbook()
    ws_matched = wb_matched.active
    ws_matched.append(["timestamp", "Meta_data", "SourceID", "TargetID", 
                       "Title", "size", "comment", 
                       'sourceDisplayName', 'targetDisplayName',
                       'sourcePublicName', 'targetPublicName', 
                       'sourceDate', 'targetDate'])
    
    wb_missing = openpyxl.Workbook()
    ws_missing = wb_missing.active
    ws_missing.append(["timestamp", "Meta_data", "SourceID", "TargetID", 
                       "Title", "size", "comment", 
                       'sourceDisplayName', 'targetDisplayName',
                       'sourcePublicName', 'targetPublicName', 
                       'sourceDate', 'targetDate'])

    # Call the initiate_space_migration function for the specific space
    initiate_space_migration(space_key, ws_matched, ws_missing)

    # Save the XLSX files for this space
    # with wb_matched_lock:
        # wb_matched.save(f"matched_attachments_{space_key}.xlsx")
    xlsx_queue.put(f"matched_attachments_{space_key}.xlsx")
    # with wb_missing_lock:
        # wb_missing.save(f"missing_attachments_{space_key}.xlsx")
    xlsx_queue.put(f"missing_attachments_{space_key}.xlsx")
    return [list_matchedData, list_missingData]
def get_spaces():
    api_endpoint = f"{target_base_url}/rest/api/space"
    response = requests.get(api_endpoint, headers=headers, auth=(username, password))
    space_results = response.json()['results']
    threads = []

    for space in space_results:
        space_key = space['key']
        # Create a thread for each space
        thread = threading.Thread(target=process_space, args=(space_key,))
        threads.append(thread)
        thread.start()

    # Wait for all threads to finish
    for thread in threads:
        thread.join()

    xlsx_files = []
    while not xlsx_queue.empty():
        xlsx_files.append(xlsx_queue.get())

    return xlsx_files

# lis_xlsx_files = get_spaces()
# print("List of XLSX files:", lis_xlsx_files)


